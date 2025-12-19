import json
from openpyxl import load_workbook

filePath_MarketStack    = r"filePath\DailyClosing_IBM_Marketstack.json"

opened_MarketStack = open(filePath_MarketStack, "r", encoding="utf-8")
key_MarketStack = "data"
data_MarketStack = json.load(opened_MarketStack)[key_MarketStack]


wb = load_workbook(r"filePath\overviewData.xlsx")
ws = wb["Overview Stock-data"]

for row in range(2, ws.max_row + 1):
    date_row = ws.cell(row, 1).value

    for i in range(len(data_MarketStack)):
        if date_row == data_MarketStack[i]["date"][0:10]:
            
            ws.cell(row, 4).value = data_MarketStack[i]["adj_open"]
            ws.cell(row, 5).value = data_MarketStack[i]["open"]

            ws.cell(row, 10).value = data_MarketStack[i]["adj_close"]
            ws.cell(row, 11).value = data_MarketStack[i]["close"]

            ws.cell(row, 16).value = data_MarketStack[i]["adj_volume"]
            ws.cell(row, 17).value = data_MarketStack[i]["volume"]
            break
            

wb.save(r"filePath\overviewData.xlsx")

opened_MarketStack.close